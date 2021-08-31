from selenium import webdriver
import openpyxl as op
import xlsxwriter as xl
from selenium.common.exceptions import NoSuchElementException
from xlsxwriter import Workbook

wb = op.load_workbook('Roll_No.xlsx')  # To open Roll Numbers Sheet
Sheet = wb['Sheet1']

subjects = []        # List for storing Subjects
count_dict1 = {}     # Dictionary for storing A+ Grades in each Subjects
count_dict2 = {}     # Dictionary for storing A Grades in each Subjects
Classes = []         # List for storing All Classes Names
colNum = 0



def find_sub(code):
    print("In find_sub function")

    for i in Sheet[1]:                      # To find the Total Classes
        if i.value:
            Classes.append(i.value)
    print(Classes)
    print(code)
    colNum = Classes.index(code) + 1        # To find column number
    print(colNum)

    driver = webdriver.Chrome()
    driver.get("https://sis.vce.ac.in/BE_Results_367_Sem_27-03-2021/")
    searchBox = driver.find_element_by_xpath('//*[@id="txtHTNO"]')    # //*[@id="lblCGPA"]
    searchButton = driver.find_element_by_xpath('//*[@id="btnResults"]')
    roll = Sheet.cell(row=2, column=colNum).value
    searchBox.send_keys(roll)
    searchButton.click()

    def check_exists_by_xpath(xpath):           # To check xpath exists or not
        try:
            driver.find_element_by_xpath(xpath)
        except NoSuchElementException:
            return False
        return True

    if check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[17]/td[1]'):
        num = 15
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[16]/td[1]'):
        num = 14
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[15]/td[1]'):
        num = 13
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[14]/td[1]'):
        num = 12
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[13]/td[1]'):
        num = 11
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[12]/td[1]'):
        num = 10
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[11]/td[1]'):
        num = 9
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[10]/td[1]'):
        num = 8
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[9]/td[1]'):
        num = 7
    elif check_exists_by_xpath('//*[@id="AutoNumber3"]/tbody/tr[8]/td[1]'):
        num = 6
    else:
        num = 5

    st = '(//*[@id="AutoNumber3"]/tbody/tr[3]/td[3])[2]'    # Manually taking Subject 1&2 Names
    sub_box = driver.find_element_by_xpath(st)
    sub_name = sub_box.text
    subjects.append(sub_name)
    st = '(//*[@id="AutoNumber3"]/tbody/tr[4]/td[3])[2]'
    sub_box = driver.find_element_by_xpath(st)
    sub_name = sub_box.text
    subjects.append(sub_name)

    for i in range(5, num + 3):                             # For getting Subject Names
        v = str(i)
        st = '(//*[@id="AutoNumber3"]/tbody/tr[' + v + ']/td[3])'
        sub_box = driver.find_element_by_xpath(st)
        sub_name = sub_box.text
        subjects.append(sub_name)

    print(subjects, "\n\n")
    driver.close()


def collect(code, filename):
    print('In collect function')
    List_Roll = []                               # List to store Roll Numbers
    file = str(filename) + '.xlsx'               # Storing xlsx file name as String
    workbook: Workbook = xl.Workbook(file)       # Creating a xlsx File
    ws = workbook.add_worksheet()                # Creating a WorkSheet
    for i in Sheet[1]:                           # For Column Number
        if i.value:
            Classes.append(i.value)
    colNum = Classes.index(code) + 1
    print(colNum)

    i = 0
    while Sheet.cell(row=i + 1, column=colNum).value:  # collects and stores roll numbers in List based on colNum
        Roll_No = Sheet.cell(row=i + 1, column=colNum).value
        List_Roll.append(Roll_No)
        i = i + 1
    class_length = len(List_Roll)                 # To find class length to Iterate
    print(class_length)

    for a in range(len(subjects)):                # To initialize dictionary key(Subject_Name):Value(0)
        count_dict1[subjects[a]] = 0              # For A+ Grades
        count_dict2[subjects[a]] = 0              # For A Grades

    driver = webdriver.Chrome()
    driver.get("https://sis.vce.ac.in/BE_Results_367_Sem_27-03-2021/")

    ws.write(0, 0, 'Roll Number')
    ws.write(0, 1, 'Name')
    ws.write(0, 2, 'SGPA')
    ws.write(0, 3, 'CGPA')

    for i in range(1, class_length):  # For Iterating through number of students

        searchBox = driver.find_element_by_xpath('//*[@id="txtHTNO"]')
        searchButton = driver.find_element_by_xpath('//*[@id="btnResults"]')

        Roll_No = List_Roll.__getitem__(i)
        searchBox.send_keys(Roll_No)
        searchButton.click()

        name = driver.find_element_by_xpath('//*[@id="lblStudName"]').text
        SGPA = driver.find_element_by_xpath('//*[@id="lblSGPa"]').text
        CGPA = driver.find_element_by_xpath('//*[@id="lblCGPA"]').text

        lis = list(SGPA.split(" "))  # For last string i.e., SGPA
        lis2 = list(CGPA.split(" "))  # For last string i.e., CGPA
        if lis[0] == '':
            sgpa = 0
        else:
            sgpa = float(lis[6])

        if lis2[0] == '':
            cgpa = 0
        else:
            cgpa = float(lis2[6])

        for j in range(len(subjects)):              # For counting A+ & A Grades in each subjects
            v = str(j + 3)
            gd = '(//*[@id="AutoNumber3"]/tbody/tr[' + v + ']/td[7])'
            grade_point = driver.find_element_by_xpath(gd).text
            if grade_point == "A+":
                count_dict1[subjects[j]] = count_dict1[subjects[j]] + 1
            if grade_point == "A":
                count_dict2[subjects[j]] = count_dict2[subjects[j]] + 1


        # For writing data into sheet
        ws.write(i, 0, Roll_No)
        ws.write(i, 1, name)
        ws.write(i, 2, sgpa)
        ws.write(i, 3, cgpa)

        # For Printing Details of each student
        print(Roll_No)
        print(name)
        print("SGPA : ", sgpa)
        print("CGPA : ", cgpa)
        print(count_dict1)
        print(count_dict2, "\n\n")

        driver.back()
        driver.refresh()

    # print('After Iterating N students Times')

    def sub_grade():        # For write number of A+ & A Grades in each Subject into sheet
        length = class_length + 10

        ws.write(length, 0, 'Subjects')
        ws.write(length, 1, 'Number of A+ Grades')
        ws.write(length, 2, 'Number of A Grades')

        for k in range(len(subjects)):
            ws.write(k + length + 1, 0, subjects[k])
            ws.write(k + length + 1, 1, count_dict1[subjects[k]])
            ws.write(k + length + 1, 2, count_dict2[subjects[k]])


    sub_grade()

    workbook.close()
    driver.close()
